from __future__ import annotations

import hashlib
import re
import unicodedata
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def _normalizar_texto(value: Any) -> str:
    texto = str(value or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "_", texto).strip("_")


def _normalizar_coluna(value: Any) -> str:
    return _normalizar_texto(value)


COLUNAS: dict[str, str] = {
    "empresa": "nome",
    "nome": "nome",
    "nome_da_empresa": "nome",
    "telefone": "telefone",
    "telefone_limpo": "telefone_limpo",
    "whatsapp_status": "whatsapp_status",
    "endereco": "endereco",
    "google_maps": "google_maps_url",
    "google_maps_url": "google_maps_url",
    "link_google_maps": "google_maps_url",
    "avaliacao": "avaliacao",
    "quantidade_avaliacoes": "quantidade_avaliacoes",
    "qtd_avaliacoes": "quantidade_avaliacoes",
    "site": "site_cadastrado",
    "site_cadastrado": "site_cadastrado",
    "websiteuri": "site_cadastrado",
    "website_uri": "site_cadastrado",
    "sem_site": "sem_site_cadastrado",
    "sem_site_cadastrado": "sem_site_cadastrado",
    "score": "score_oportunidade",
    "score_oportunidade": "score_oportunidade",
    "classificacao": "classificacao_lead",
    "classificacao_lead": "classificacao_lead",
    "prioridade": "prioridade",
    "regiao": "regiao",
    "cidade": "cidade",
    "segmento": "segmento",
    "oferta_principal": "oferta_principal",
    "observacao_comercial": "observacao_comercial",
    "status": "status_contato",
    "status_contato": "status_contato",
    "status_do_contato": "status_contato",
    "etapa": "status_contato",
    "etapa_funil": "status_contato",
    "data_primeiro_contato": "data_primeiro_contato",
    "primeiro_contato": "data_primeiro_contato",
    "data_ultimo_contato": "data_ultimo_contato",
    "ultimo_contato": "data_ultimo_contato",
    "proximo_followup": "proximo_followup",
    "proximo_follow_up": "proximo_followup",
    "respondeu": "respondeu",
    "interesse": "interesse",
    "diagnostico_enviado": "diagnostico_enviado",
    "reuniao_marcada": "reuniao_marcada",
    "proposta_enviada": "proposta_enviada",
    "fechado": "fechado",
    "motivo_perda": "motivo_perda",
    "observacao_humana": "observacao_humana",
    "place_id": "place_id",
    "business_status": "business_status",
}


STATUS_ALIASES: dict[str, str] = {
    "novo": "Novo",
    "nao_contatado": "Novo",
    "sem_contato": "Novo",
    "contactado": "Primeiro contato",
    "contactada": "Primeiro contato",
    "contactados": "Primeiro contato",
    "contactadas": "Primeiro contato",
    "contatado": "Primeiro contato",
    "contatada": "Primeiro contato",
    "contatados": "Primeiro contato",
    "contatadas": "Primeiro contato",
    "em_contato": "Primeiro contato",
    "contato_feito": "Primeiro contato",
    "contato_realizado": "Primeiro contato",
    "primeiro_contato": "Primeiro contato",
    "1_contato": "Primeiro contato",
    "primeira_abordagem": "Primeiro contato",
    "abordado": "Primeiro contato",
    "abordada": "Primeiro contato",
    "abordagem_feita": "Primeiro contato",
    "respondeu": "Respondeu",
    "respondido": "Respondeu",
    "com_resposta": "Respondeu",
    "diagnostico": "Diagnóstico enviado",
    "diagnostico_enviado": "Diagnóstico enviado",
    "reuniao": "Reunião marcada",
    "reuniao_marcada": "Reunião marcada",
    "proposta": "Proposta",
    "proposta_enviada": "Proposta",
    "fechado": "Fechado",
    "cliente_fechado": "Fechado",
    "perdido": "Perdido",
}


def normalizar_status_contato(value: Any) -> str:
    return _status(value) or "Novo"


def _texto(value: Any) -> str | None:
    if value is None:
        return None
    texto = str(value).strip()
    return texto or None


def _inteiro(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).replace(",", ".")))
    except ValueError:
        return None


def _decimal(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def _booleano(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    texto = _normalizar_texto(value)
    if texto in {"sim", "s", "yes", "true", "1"}:
        return True
    if texto in {"nao", "n", "no", "false", "0"}:
        return False
    return None


def _data(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    texto = str(value).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def _status(value: Any) -> str | None:
    texto = _texto(value)
    if not texto:
        return None
    return STATUS_ALIASES.get(_normalizar_texto(texto), texto)


def _somente_numeros(value: str | None) -> str:
    return re.sub(r"\D", "", value or "")


def _telefone_limpo(telefone: str | None, telefone_limpo: str | None) -> str | None:
    numero = _somente_numeros(telefone_limpo or telefone)
    if not numero:
        return None
    if numero.startswith("55"):
        return numero
    if 10 <= len(numero) <= 11:
        return f"55{numero}"
    return numero


def _whatsapp_status(telefone_limpo: str | None, informado: str | None) -> str:
    if informado:
        return informado
    if not telefone_limpo:
        return "Sem telefone"
    if re.match(r"^55[1-9][0-9]9[0-9]{8}$", telefone_limpo):
        return "Provável WhatsApp"
    return "Verificar"


def _place_id(row: dict[str, Any]) -> str:
    existente = _texto(row.get("place_id"))
    if existente:
        return existente
    base = "|".join(
        [
            _texto(row.get("google_maps_url")) or "",
            _texto(row.get("nome")) or "",
            _texto(row.get("cidade")) or "",
            _texto(row.get("telefone_limpo")) or _texto(row.get("telefone")) or "",
        ]
    ).lower()
    return f"import:{hashlib.sha1(base.encode('utf-8')).hexdigest()}"


def _inferir_status(cleaned: dict[str, Any]) -> str:
    status = _status(cleaned.get("status_contato"))
    if status:
        return status
    if cleaned.get("fechado"):
        return "Fechado"
    if cleaned.get("proposta_enviada"):
        return "Proposta"
    if cleaned.get("reuniao_marcada"):
        return "Reunião marcada"
    if cleaned.get("diagnostico_enviado"):
        return "Diagnóstico enviado"
    if cleaned.get("respondeu"):
        return "Respondeu"
    if cleaned.get("data_primeiro_contato") or cleaned.get("data_ultimo_contato"):
        return "Primeiro contato"
    return "Novo"


def _limpar_row(row: dict[str, Any]) -> dict[str, Any]:
    texto_fields = {
        "place_id",
        "nome",
        "telefone",
        "telefone_limpo",
        "whatsapp_status",
        "endereco",
        "cidade",
        "segmento",
        "regiao",
        "google_maps_url",
        "site_cadastrado",
        "sem_site_cadastrado",
        "business_status",
        "classificacao_lead",
        "prioridade",
        "oferta_principal",
        "observacao_comercial",
        "interesse",
        "motivo_perda",
        "observacao_humana",
    }
    int_fields = {"quantidade_avaliacoes", "score_oportunidade"}
    decimal_fields = {"avaliacao"}
    bool_fields = {"respondeu", "diagnostico_enviado", "reuniao_marcada", "proposta_enviada", "fechado"}
    date_fields = {"data_primeiro_contato", "data_ultimo_contato", "proximo_followup"}

    cleaned: dict[str, Any] = {}
    for field, value in row.items():
        if field in texto_fields:
            cleaned[field] = _texto(value)
        elif field == "status_contato":
            cleaned[field] = _status(value)
        elif field in int_fields:
            cleaned[field] = _inteiro(value)
        elif field in decimal_fields:
            cleaned[field] = _decimal(value)
        elif field in bool_fields:
            cleaned[field] = _booleano(value)
        elif field in date_fields:
            cleaned[field] = _data(value)
        else:
            cleaned[field] = value

    cleaned["telefone_limpo"] = _telefone_limpo(cleaned.get("telefone"), cleaned.get("telefone_limpo"))
    cleaned["whatsapp_status"] = _whatsapp_status(cleaned.get("telefone_limpo"), cleaned.get("whatsapp_status"))
    cleaned["sem_site_cadastrado"] = cleaned.get("sem_site_cadastrado") or (
        "SIM" if not cleaned.get("site_cadastrado") else "NÃO"
    )
    cleaned["status_contato"] = _inferir_status(cleaned)
    cleaned["place_id"] = _place_id(cleaned)
    return cleaned


def extrair_leads_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    mapped_headers = [COLUNAS.get(_normalizar_coluna(header)) for header in headers]
    leads: list[dict[str, Any]] = []

    for raw_row in rows:
        row: dict[str, Any] = {}
        for index, value in enumerate(raw_row):
            field = mapped_headers[index] if index < len(mapped_headers) else None
            if field and value not in (None, ""):
                row[field] = value
        cleaned = _limpar_row(row)
        if cleaned.get("nome"):
            leads.append(cleaned)

    return leads

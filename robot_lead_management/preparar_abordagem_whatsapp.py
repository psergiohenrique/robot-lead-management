"""Prepara uma planilha para abordagem manual de leads pelo WhatsApp.

O script não envia mensagens. Ele apenas organiza a base de empresas sem site,
cria textos personalizados e gera links ``wa.me`` para abertura manual.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path.cwd()
OUTPUT_DIR = BASE_DIR / "output"
ARQUIVO_FINAL = OUTPUT_DIR / "abordagem_whatsapp_sem_site.xlsx"
CODEPATH_SITE_PROMOCIONAL = "https://promocao.codepath.dev.br/"
CODEPATH_INSTAGRAM = "@codepath.softwares"

COLUNAS_GERADAS = [
    "telefone_limpo",
    "whatsapp_status",
    "link_whatsapp",
    "abordagem_tipo",
    "mensagem_whatsapp",
    "status_contato",
    "data_primeiro_contato",
    "data_ultimo_contato",
    "proximo_followup",
    "respondeu",
    "interesse",
    "diagnostico_enviado",
    "reuniao_marcada",
    "proposta_enviada",
    "fechado",
    "motivo_perda",
    "observacao_humana",
    "sugestao_validacao_manual",
]

COLUNAS_CONTROLE = {
    "status_contato",
    "data_primeiro_contato",
    "data_ultimo_contato",
    "proximo_followup",
    "respondeu",
    "interesse",
    "diagnostico_enviado",
    "reuniao_marcada",
    "proposta_enviada",
    "fechado",
    "motivo_perda",
    "observacao_humana",
}

ALIASES = {
    "empresa": ["empresa", "nome", "nome da empresa", "razao social"],
    "cidade": ["cidade", "municipio"],
    "segmento": ["segmento", "categoria", "ramo"],
    "telefone": ["telefone", "telefone principal", "phone"],
    "endereco": ["endereco", "endereço"],
    "google_maps": ["google maps", "google_maps", "link google maps", "maps"],
    "avaliacao": ["avaliacao", "avaliação", "nota"],
    "quantidade_avaliacoes": [
        "quantidade de avaliacoes",
        "quantidade_avaliacoes",
        "numero de avaliacoes",
    ],
    "sem_site": ["sem site cadastrado", "sem_site_cadastrado", "sem site"],
    "score": ["score", "score oportunidade", "score_oportunidade"],
    "classificacao": [
        "classificacao",
        "classificação",
        "classificacao lead",
        "classificacao_lead",
    ],
    "prioridade": ["prioridade"],
    "regiao": ["regiao", "região"],
    "oferta": ["oferta principal", "oferta_principal"],
    "observacao": ["observacao comercial", "observação comercial", "observacao_comercial"],
    "place_id": ["place id", "place_id"],
}

COR_AZUL = "1F4E78"
COR_VERDE = "0F766E"
COR_DOURADO = "A16207"
COR_AZUL_CLARO = "DCE6F1"
COR_VERDE_CLARO = "DCFCE7"
COR_AMARELO_CLARO = "FEF3C7"
COR_VERMELHO_CLARO = "FEE2E2"
COR_CINZA_CLARO = "F3F4F6"
COR_BRANCO = "FFFFFF"


def normalizar_texto(valor: Any) -> str:
    texto = "" if valor is None else str(valor).strip()
    texto = "".join(
        caractere
        for caractere in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(caractere)
    )
    return " ".join(texto.casefold().replace("_", " ").split())


def nome_coluna_normalizado(valor: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", normalizar_texto(valor))


def valor_texto(valor: Any) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def valor_numero(valor: Any, padrao: float = 0) -> float:
    if valor in (None, ""):
        return padrao
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace(" ", "")
    if "," in texto and "." not in texto:
        texto = texto.replace(",", ".")
    else:
        texto = texto.replace(",", "")
    try:
        return float(texto)
    except ValueError:
        return padrao


def argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Gera uma planilha para abordagem manual via WhatsApp usando apenas "
            "empresas sem site cadastrado. Nenhuma mensagem é enviada."
        )
    )
    parser.add_argument(
        "--arquivo",
        help=(
            "Base .xlsx ou .csv. Se omitido, procura automaticamente em output "
            "e usa a base sem site com mais registros."
        ),
    )
    parser.add_argument(
        "--saida",
        default=str(ARQUIVO_FINAL),
        help="Caminho do Excel final.",
    )
    return parser.parse_args()


def detectar_dialeto(caminho: Path) -> csv.Dialect:
    with caminho.open("r", encoding="utf-8-sig", newline="") as arquivo:
        amostra = arquivo.read(8192)
    try:
        return csv.Sniffer().sniff(amostra, delimiters=";,\t")
    except csv.Error:
        return csv.excel


def analisar_xlsx(caminho: Path) -> tuple[int, str, list[str]]:
    workbook = load_workbook(caminho, read_only=True, data_only=True)
    melhor: tuple[int, str, list[str]] | None = None
    try:
        for planilha in workbook.worksheets:
            linhas = planilha.iter_rows(values_only=True)
            cabecalho_bruto = next(linhas, None)
            if not cabecalho_bruto:
                continue
            cabecalho = [valor_texto(item) for item in cabecalho_bruto]
            if not any(cabecalho):
                continue
            quantidade = sum(
                1 for linha in linhas if any(item not in (None, "") for item in linha)
            )
            candidato = (quantidade, planilha.title, cabecalho)
            if melhor is None or quantidade > melhor[0]:
                melhor = candidato
    finally:
        workbook.close()
    if melhor is None:
        raise ValueError("Nenhuma aba com dados foi encontrada.")
    return melhor


def analisar_csv(caminho: Path) -> tuple[int, str, list[str]]:
    dialeto = detectar_dialeto(caminho)
    with caminho.open("r", encoding="utf-8-sig", newline="") as arquivo:
        leitor = csv.reader(arquivo, dialect=dialeto)
        cabecalho = [valor_texto(item) for item in (next(leitor, None) or [])]
        quantidade = sum(1 for linha in leitor if any(valor_texto(item) for item in linha))
    if not cabecalho:
        raise ValueError("O arquivo não possui cabeçalho.")
    return quantidade, "CSV", cabecalho


def analisar_arquivo(caminho: Path) -> tuple[int, str, list[str]]:
    if caminho.suffix.lower() == ".xlsx":
        return analisar_xlsx(caminho)
    if caminho.suffix.lower() == ".csv":
        return analisar_csv(caminho)
    raise ValueError("Formato não aceito. Use .xlsx ou .csv.")


def descobrir_arquivos_sem_site() -> list[dict[str, Any]]:
    encontrados: list[dict[str, Any]] = []
    if not OUTPUT_DIR.exists():
        return encontrados
    for caminho in sorted(OUTPUT_DIR.rglob("*")):
        if not caminho.is_file() or caminho.suffix.lower() not in {".xlsx", ".csv"}:
            continue
        if caminho.resolve() == ARQUIVO_FINAL.resolve():
            continue
        if "semsite" not in nome_coluna_normalizado(caminho.stem):
            continue
        try:
            quantidade, aba, cabecalho = analisar_arquivo(caminho)
        except (OSError, ValueError, zipfile.BadZipFile) as erro:
            print(f"  Ignorado: {caminho} ({erro})")
            continue
        encontrados.append(
            {
                "caminho": caminho,
                "registros": quantidade,
                "aba": aba,
                "cabecalho": cabecalho,
                "modificado": caminho.stat().st_mtime,
            }
        )
    return encontrados


def escolher_arquivo(caminho_informado: str | None) -> tuple[Path, str, list[str]]:
    if caminho_informado:
        caminho = Path(caminho_informado)
        if not caminho.is_absolute():
            caminho = BASE_DIR / caminho
        if not caminho.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")
        quantidade, aba, cabecalho = analisar_arquivo(caminho)
        print(f"Base informada: {caminho} ({quantidade} registros)")
        return caminho, aba, cabecalho

    encontrados = descobrir_arquivos_sem_site()
    if not encontrados:
        raise FileNotFoundError(
            "Nenhuma planilha de empresas sem site foi encontrada dentro de output."
        )

    print("Arquivos de empresas sem site encontrados:")
    for item in sorted(
        encontrados, key=lambda dado: (-dado["registros"], -dado["modificado"])
    ):
        relativo = item["caminho"].relative_to(BASE_DIR)
        print(f"  - {relativo}: {item['registros']} registros")

    escolhido = max(
        encontrados,
        key=lambda dado: (dado["registros"], dado["modificado"], dado["caminho"].suffix == ".xlsx"),
    )
    print(
        "Base escolhida automaticamente: "
        f"{escolhido['caminho'].relative_to(BASE_DIR)} "
        f"({escolhido['registros']} registros)"
    )
    return escolhido["caminho"], escolhido["aba"], escolhido["cabecalho"]


def ler_xlsx(caminho: Path, nome_aba: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(caminho, read_only=True, data_only=True)
    try:
        planilha = workbook[nome_aba]
        linhas = planilha.iter_rows(values_only=True)
        cabecalho = [valor_texto(item) for item in (next(linhas, None) or [])]
        registros: list[dict[str, Any]] = []
        for linha in linhas:
            if not any(item not in (None, "") for item in linha):
                continue
            registros.append(
                {
                    coluna: linha[indice] if indice < len(linha) else None
                    for indice, coluna in enumerate(cabecalho)
                    if coluna
                }
            )
        return [coluna for coluna in cabecalho if coluna], registros
    finally:
        workbook.close()


def ler_csv(caminho: Path) -> tuple[list[str], list[dict[str, Any]]]:
    dialeto = detectar_dialeto(caminho)
    with caminho.open("r", encoding="utf-8-sig", newline="") as arquivo:
        leitor = csv.DictReader(arquivo, dialect=dialeto)
        cabecalho = [valor_texto(item) for item in (leitor.fieldnames or []) if item]
        registros = [dict(linha) for linha in leitor if any(valor_texto(v) for v in linha.values())]
    return cabecalho, registros


def ler_dados(caminho: Path, nome_aba: str) -> tuple[list[str], list[dict[str, Any]]]:
    if caminho.suffix.lower() == ".xlsx":
        return ler_xlsx(caminho, nome_aba)
    return ler_csv(caminho)


def mapear_colunas(cabecalho: Iterable[str]) -> dict[str, str | None]:
    por_nome = {nome_coluna_normalizado(coluna): coluna for coluna in cabecalho}
    mapa: dict[str, str | None] = {}
    for campo, aliases in ALIASES.items():
        mapa[campo] = next(
            (por_nome[nome_coluna_normalizado(alias)] for alias in aliases if nome_coluna_normalizado(alias) in por_nome),
            None,
        )
    return mapa


def obter(registro: dict[str, Any], mapa: dict[str, str | None], campo: str) -> Any:
    coluna = mapa.get(campo)
    return registro.get(coluna) if coluna else None


def limpar_telefone(valor: Any) -> str:
    digitos = re.sub(r"\D", "", valor_texto(valor))
    if digitos.startswith("00"):
        digitos = digitos[2:]
    if not digitos:
        return ""
    if digitos.startswith("55") and len(digitos) in {12, 13}:
        return digitos
    return f"55{digitos}"


def analisar_telefone(telefone_limpo: str) -> tuple[str, bool]:
    if not telefone_limpo:
        return "Sem telefone", False
    nacional = telefone_limpo[2:] if telefone_limpo.startswith("55") else telefone_limpo
    valido_com_ddd = len(nacional) in {10, 11} and nacional[:2].isdigit() and nacional[:2] != "00"
    if valido_com_ddd and len(nacional) == 11 and nacional[2] == "9":
        return "Provável WhatsApp", True
    if valido_com_ddd:
        return "Verificar", True
    if len(nacional) < 8:
        return "Sem telefone", False
    return "Verificar", False


def classificar_abordagem(segmento: Any) -> str:
    texto = normalizar_texto(segmento)

    # Medicina do trabalho aparece nas regras de Saúde e B2B. Como geralmente é
    # contratada por empresas, adotamos B2B como classificação principal.
    termos_b2b = [
        "medicina do trabalho",
        "pequena industria",
        "consultoria empresarial",
        "servicos empresariais",
        "logistica",
        "industria",
    ]
    termos_saude = [
        "dentista",
        "odontolog",
        "clinica medica",
        "dermatolog",
        "cardiolog",
        "ortoped",
        "ginecolog",
        "pediatr",
        "oftalmolog",
        "endocrinolog",
        "urolog",
        "psiquiatr",
        "neurolog",
        "clinica estetica",
        "cirurgiao plastico",
        "cirurgia plastica",
        "emagrecimento",
    ]
    termos_premium = [
        "arquitetura",
        "moveis planejados",
        "imobiliaria",
        "decoracao",
        "construtora",
        "construcao de alto valor",
    ]
    termos_local = [
        "comercio",
        "restaurante",
        "oficina",
        "auto center",
        "academia",
        "veterin",
        "escola",
        "curso",
        "loja",
        "hotel",
        "pousada",
        "servico local",
    ]

    if any(termo in texto for termo in termos_b2b):
        return "B2B"
    if any(termo in texto for termo in termos_saude):
        return "Saúde"
    if any(termo in texto for termo in termos_premium):
        return "Premium/Portfólio"
    if any(termo in texto for termo in termos_local):
        return "Local"
    return "Geral"


def cidade_para_mensagem(valor: Any) -> str:
    cidade = valor_texto(valor)
    cidade = re.sub(r"\s+(?:SP|RJ|MG|PR|SC|RS|BA|GO|DF|ES|MS|MT|PE|CE)$", "", cidade, flags=re.I)
    return cidade or "sua região"


def indice_variacao(registro: dict[str, Any], mapa: dict[str, str | None]) -> int:
    chave = valor_texto(obter(registro, mapa, "place_id")) or valor_texto(
        obter(registro, mapa, "empresa")
    )
    return sum(ord(caractere) for caractere in chave) % 2


def criar_mensagem(
    abordagem: str,
    cidade: str,
    variacao: int,
) -> str:
    saudacoes = ["Olá, tudo bem?", "Olá! Tudo bem por aí?"]
    motivos = {
        "Saúde": (
            "Na área da saúde, isso pode influenciar a confiança do paciente antes do "
            "primeiro contato."
        ),
        "B2B": (
            "Para empresas B2B, um site ajuda a apresentar a empresa, serviços e canais "
            "de contato com mais credibilidade para clientes e parceiros."
        ),
        "Premium/Portfólio": (
            "Em serviços de alto valor, um site ajuda a apresentar portfólio, diferenciais "
            "e formas de contato com mais clareza."
        ),
        "Local": (
            "Isso pode fazer alguns clientes terem menos informações antes de chamar no "
            "WhatsApp ou solicitar orçamento."
        ),
        "Geral": (
            "Isso pode deixar informações importantes menos acessíveis antes do primeiro "
            "contato."
        ),
    }
    chamadas = [
        (
            "Posso te enviar mais detalhes dessa condição e, se fizer sentido, um "
            "*diagnóstico rápido* do perfil de vocês?"
        ),
        (
            "Posso te mandar os detalhes da promoção e uma análise rápida do que poderia "
            "melhorar no perfil de vocês?"
        ),
    ]
    promocao = (
        "*Promoção Codepath*\n"
        "*Site institucional completo*\n"
        "*R$ 499 à vista*\n"
        "+ *R$ 129,90/mês* de manutenção, suporte e cuidados contínuos do site enquanto a Codepath cuidar dele\n\n"
        "Inclui:\n"
        "- Site profissional\n"
        "- Layout responsivo\n"
        "- Estrutura pensada para o Google\n"
        "- Suporte direto da Codepath\n"
        "- *Vagas limitadas nessa condição promocional*"
    )
    canais = (
        "Para conhecer melhor a *Codepath*:\n"
        f"Site: {CODEPATH_SITE_PROMOCIONAL}\n"
        f"Instagram: {CODEPATH_INSTAGRAM}"
    )

    return (
        f"{saudacoes[variacao % len(saudacoes)]}\n\n"
        "Sou da *Codepath*. Estamos com uma *condição promocional por tempo limitado* "
        "para criação de *site institucional completo*.\n\n"
        f"{promocao}\n\n"
        f"Vi que vocês aparecem no Google em {cidade}, mas não encontrei um "
        f"*site cadastrado no perfil*.\n\n"
        f"{motivos.get(abordagem, motivos['Geral'])}\n\n"
        f"{chamadas[variacao % len(chamadas)]}\n\n"
        f"{canais}"
    )


def classificacao_rank(valor: Any) -> int:
    texto = normalizar_texto(valor)
    if "quente" in texto:
        return 3
    if "morno" in texto:
        return 2
    if "frio" in texto:
        return 1
    return 0


def prioridade_rank(valor: Any) -> int:
    texto = normalizar_texto(valor)
    if texto == "alta":
        return 2
    if texto == "media":
        return 1
    return 0


def whatsapp_rank(valor: Any) -> int:
    return {"Provável WhatsApp": 3, "Verificar": 2, "Sem telefone": 1}.get(valor_texto(valor), 0)


def sugestao_validacao(
    prioridade: Any,
    status: str,
    score: float,
    classificacao: Any,
) -> str:
    sugestoes: list[str] = []
    if normalizar_texto(prioridade) == "alta":
        sugestoes.append("Validar Instagram e Google antes do envio")
    if status == "Verificar":
        sugestoes.append("Confirmar se telefone é WhatsApp")
    if status == "Provável WhatsApp" and score >= 8 and classificacao_rank(classificacao) == 3:
        sugestoes.append("Lead pronto para abordagem")
    if score <= 4 or classificacao_rank(classificacao) == 1:
        sugestoes.append("Baixa prioridade inicial")
    if status == "Sem telefone":
        sugestoes.append("Buscar outro canal de contato")
    return " | ".join(sugestoes) or "Revisar perfil antes do envio"


def preparar_registros(
    registros: list[dict[str, Any]],
    mapa: dict[str, str | None],
) -> list[dict[str, Any]]:
    preparados: list[dict[str, Any]] = []
    for original in registros:
        sem_site = normalizar_texto(obter(original, mapa, "sem_site"))
        if mapa.get("sem_site") and sem_site not in {"sim", "yes", "true", "1"}:
            continue

        registro = dict(original)
        telefone_limpo = limpar_telefone(obter(original, mapa, "telefone"))
        status, telefone_valido = analisar_telefone(telefone_limpo)
        abordagem = classificar_abordagem(obter(original, mapa, "segmento"))
        cidade = cidade_para_mensagem(obter(original, mapa, "cidade"))
        mensagem = criar_mensagem(abordagem, cidade, indice_variacao(original, mapa))
        score = valor_numero(obter(original, mapa, "score"))
        classificacao = obter(original, mapa, "classificacao")

        registro.update(
            {
                "telefone_limpo": telefone_limpo,
                "whatsapp_status": status,
                "link_whatsapp": (
                    f"https://wa.me/{telefone_limpo}?text={quote(mensagem, safe='')}"
                    if telefone_valido
                    else ""
                ),
                "abordagem_tipo": abordagem,
                "mensagem_whatsapp": mensagem,
                "status_contato": "Novo",
                "data_primeiro_contato": "",
                "data_ultimo_contato": "",
                "proximo_followup": "",
                "respondeu": "Não",
                "interesse": "",
                "diagnostico_enviado": "Não",
                "reuniao_marcada": "Não",
                "proposta_enviada": "Não",
                "fechado": "Não",
                "motivo_perda": "",
                "observacao_humana": "",
                "sugestao_validacao_manual": sugestao_validacao(
                    obter(original, mapa, "prioridade"), status, score, classificacao
                ),
                "_telefone_valido": telefone_valido,
                "_score_numero": score,
                "_classificacao_rank": classificacao_rank(classificacao),
                "_prioridade_rank": prioridade_rank(obter(original, mapa, "prioridade")),
                "_avaliacoes_numero": valor_numero(
                    obter(original, mapa, "quantidade_avaliacoes")
                ),
            }
        )
        preparados.append(registro)
    return preparados


def ordenar_base(
    registros: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    return sorted(
        registros,
        key=lambda item: (
            -item["_classificacao_rank"],
            -item["_score_numero"],
            -whatsapp_rank(item["whatsapp_status"]),
            -item["_avaliacoes_numero"],
            -item["_prioridade_rank"],
        ),
    )


def selecionar_top_50(
    registros: list[dict[str, Any]],
    mapa: dict[str, str | None],
) -> list[dict[str, Any]]:
    rank_abordagem = {"Saúde": 4, "B2B": 3, "Premium/Portfólio": 2, "Local": 1, "Geral": 0}
    cidades_centrais = {"sao jose do rio preto sp", "campinas sp"}

    elegiveis = [
        item
        for item in registros
        if item["_telefone_valido"] and item["whatsapp_status"] == "Provável WhatsApp"
    ]
    elegiveis.sort(
        key=lambda item: (
            -item["_classificacao_rank"],
            -item["_score_numero"],
            -item["_avaliacoes_numero"],
            -item["_prioridade_rank"],
            -rank_abordagem.get(item["abordagem_tipo"], 0),
            -(normalizar_texto(obter(item, mapa, "cidade")) in cidades_centrais),
            -bool(valor_texto(obter(item, mapa, "regiao"))),
        )
    )
    return elegiveis[:50]


def limpar_valor_excel(valor: Any) -> Any:
    if isinstance(valor, str):
        return ILLEGAL_CHARACTERS_RE.sub("", valor)
    return valor


def largura_coluna(nome: str) -> float:
    normal = nome_coluna_normalizado(nome)
    if normal in {"placeid", "telefonelimpo", "telefone"}:
        return 20
    if normal in {"nome", "empresa", "nomedaempresa"}:
        return 38
    if normal in {"endereco", "observacaocomercial", "motivooportunidade", "observacaohumana"}:
        return 46
    if normal in {"googlemaps", "linkwhatsapp"}:
        return 36
    if normal == "mensagemwhatsapp":
        return 85
    if normal in {"sugestaovalidacaomanual", "consultasencontradas", "querybase"}:
        return 46
    if normal.startswith("data") or normal == "proximofollowup":
        return 18
    if normal in {"segmento", "ofertaprincipal"}:
        return 28
    if normal in {"cidade", "regiao", "abordagemtipo", "whatsappstatus", "statuscontato"}:
        return 24
    return 18


def adicionar_validacoes(planilha: Any, cabecalho: list[str], ultima_linha: int) -> None:
    if ultima_linha < 2:
        return
    indice = {nome: posicao + 1 for posicao, nome in enumerate(cabecalho)}
    validacoes = {
        "status_contato": "Novo,Contatado,Respondeu,Diagnóstico enviado,Reunião marcada,Proposta enviada,Fechado,Perdido",
        "respondeu": "Não,Sim",
        "diagnostico_enviado": "Não,Sim",
        "reuniao_marcada": "Não,Sim",
        "proposta_enviada": "Não,Sim",
        "fechado": "Não,Sim",
        "interesse": "Não informado,Baixo,Médio,Alto",
    }
    for coluna, opcoes in validacoes.items():
        if coluna not in indice:
            continue
        validacao = DataValidation(type="list", formula1=f'"{opcoes}"', allow_blank=True)
        validacao.error = "Escolha uma opção da lista."
        validacao.errorTitle = "Valor inválido"
        planilha.add_data_validation(validacao)
        letra = get_column_letter(indice[coluna])
        validacao.add(f"{letra}2:{letra}{ultima_linha}")


def escrever_aba_dados(
    workbook: Workbook,
    nome: str,
    cabecalho: list[str],
    registros: list[dict[str, Any]],
    nome_tabela: str,
) -> Any:
    planilha = workbook.create_sheet(nome)
    planilha.sheet_view.showGridLines = False
    planilha.freeze_panes = "A2"
    planilha.append(cabecalho)
    for registro in registros:
        planilha.append([limpar_valor_excel(registro.get(coluna, "")) for coluna in cabecalho])

    borda = Border(bottom=Side(style="thin", color="D1D5DB"))
    originais = set(cabecalho) - set(COLUNAS_GERADAS)
    for coluna, nome_coluna in enumerate(cabecalho, start=1):
        celula = planilha.cell(1, coluna)
        if nome_coluna in COLUNAS_CONTROLE:
            cor = COR_DOURADO
        elif nome_coluna in originais:
            cor = COR_AZUL
        else:
            cor = COR_VERDE
        celula.fill = PatternFill("solid", fgColor=cor)
        celula.font = Font(color=COR_BRANCO, bold=True)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = borda
        planilha.column_dimensions[get_column_letter(coluna)].width = largura_coluna(nome_coluna)
    planilha.row_dimensions[1].height = 36

    indice = {nome_coluna: posicao + 1 for posicao, nome_coluna in enumerate(cabecalho)}
    colunas_quebra = {
        "endereco",
        "mensagem_whatsapp",
        "sugestao_validacao_manual",
        "observacao_humana",
        "observacao_comercial",
        "motivo_oportunidade",
    }
    for linha in range(2, planilha.max_row + 1):
        planilha.row_dimensions[linha].height = 75 if nome == "top_50_primeira_abordagem" else 60
        for coluna_nome in colunas_quebra:
            if coluna_nome in indice:
                planilha.cell(linha, indice[coluna_nome]).alignment = Alignment(
                    vertical="top", wrap_text=True
                )
        for coluna_nome in {"data_primeiro_contato", "data_ultimo_contato", "proximo_followup"}:
            if coluna_nome in indice:
                planilha.cell(linha, indice[coluna_nome]).number_format = "dd/mm/yyyy"
        for coluna_nome in {"telefone_limpo", "place_id"}:
            if coluna_nome in indice:
                planilha.cell(linha, indice[coluna_nome]).number_format = "@"
                planilha.cell(linha, indice[coluna_nome]).alignment = Alignment(
                    horizontal="left", vertical="top"
                )
        for coluna_nome in {"avaliacao"}:
            coluna_real = next(
                (cab for cab in cabecalho if nome_coluna_normalizado(cab) == nome_coluna_normalizado(coluna_nome)),
                None,
            )
            if coluna_real:
                planilha.cell(linha, indice[coluna_real]).number_format = "0.0"
        for coluna_nome in {"quantidade_avaliacoes", "score_oportunidade"}:
            coluna_real = next(
                (cab for cab in cabecalho if nome_coluna_normalizado(cab) == nome_coluna_normalizado(coluna_nome)),
                None,
            )
            if coluna_real:
                planilha.cell(linha, indice[coluna_real]).number_format = "#,##0"

    for coluna_link in ("link_whatsapp",):
        if coluna_link in indice:
            for linha in range(2, planilha.max_row + 1):
                celula = planilha.cell(linha, indice[coluna_link])
                if valor_texto(celula.value).startswith("https://"):
                    celula.hyperlink = celula.value
                    celula.style = "Hyperlink"

    coluna_maps = next(
        (cab for cab in cabecalho if nome_coluna_normalizado(cab) == "googlemaps"), None
    )
    if coluna_maps:
        for linha in range(2, planilha.max_row + 1):
            celula = planilha.cell(linha, indice[coluna_maps])
            if valor_texto(celula.value).startswith("https://"):
                celula.hyperlink = celula.value
                celula.style = "Hyperlink"

    if planilha.max_row >= 2:
        referencia = f"A1:{get_column_letter(len(cabecalho))}{planilha.max_row}"
        tabela = Table(displayName=nome_tabela, ref=referencia)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        planilha.add_table(tabela)

    if "whatsapp_status" in indice and planilha.max_row >= 2:
        letra = get_column_letter(indice["whatsapp_status"])
        intervalo = f"{letra}2:{letra}{planilha.max_row}"
        planilha.conditional_formatting.add(
            intervalo,
            FormulaRule(
                formula=[f'${letra}2="Provável WhatsApp"'],
                fill=PatternFill("solid", fgColor=COR_VERDE_CLARO),
            ),
        )
        planilha.conditional_formatting.add(
            intervalo,
            FormulaRule(
                formula=[f'${letra}2="Verificar"'],
                fill=PatternFill("solid", fgColor=COR_AMARELO_CLARO),
            ),
        )
        planilha.conditional_formatting.add(
            intervalo,
            FormulaRule(
                formula=[f'${letra}2="Sem telefone"'],
                fill=PatternFill("solid", fgColor=COR_VERMELHO_CLARO),
            ),
        )

    adicionar_validacoes(planilha, cabecalho, planilha.max_row)
    return planilha


def coluna_por_alias(cabecalho: list[str], aliases: Iterable[str]) -> str | None:
    procurados = {nome_coluna_normalizado(alias) for alias in aliases}
    return next((coluna for coluna in cabecalho if nome_coluna_normalizado(coluna) in procurados), None)


def criar_resumo(
    workbook: Workbook,
    cabecalho: list[str],
    registros: list[dict[str, Any]],
    mapa: dict[str, str | None],
    arquivo_origem: Path,
    top_50: list[dict[str, Any]],
) -> None:
    planilha = workbook.create_sheet("resumo")
    planilha.sheet_view.showGridLines = False
    planilha.freeze_panes = "A4"
    planilha.merge_cells("A1:H1")
    planilha["A1"] = "Resumo da abordagem manual via WhatsApp"
    planilha["A1"].fill = PatternFill("solid", fgColor=COR_AZUL)
    planilha["A1"].font = Font(color=COR_BRANCO, bold=True, size=16)
    planilha["A1"].alignment = Alignment(horizontal="center", vertical="center")
    planilha.row_dimensions[1].height = 30
    planilha.merge_cells("A2:H2")
    planilha["A2"] = (
        "Apoio ao contato individual: revise o perfil e a mensagem antes de abrir o link. "
        "Não existe envio automático."
    )
    planilha["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    planilha["A2"].fill = PatternFill("solid", fgColor=COR_AZUL_CLARO)
    planilha.row_dimensions[2].height = 36

    indice = {nome: posicao + 1 for posicao, nome in enumerate(cabecalho)}
    ultima_linha = len(registros) + 1
    letra_status = get_column_letter(indice["whatsapp_status"])
    primeira_coluna = get_column_letter(1)
    referencia_base = "'base_abordagem_completa'"

    planilha["A4"] = "Indicador"
    planilha["B4"] = "Valor"
    indicadores = [
        ("Total de leads sem site", f"=COUNTA({referencia_base}!${primeira_coluna}$2:${primeira_coluna}${ultima_linha})"),
        ("Com provável WhatsApp", f'=COUNTIF({referencia_base}!${letra_status}$2:${letra_status}${ultima_linha},"Provável WhatsApp")'),
        ("Para verificar telefone", f'=COUNTIF({referencia_base}!${letra_status}$2:${letra_status}${ultima_linha},"Verificar")'),
        ("Sem telefone", f'=COUNTIF({referencia_base}!${letra_status}$2:${letra_status}${ultima_linha},"Sem telefone")'),
        ("Selecionados para primeira abordagem", len(top_50)),
    ]
    for linha, (rotulo, valor) in enumerate(indicadores, start=5):
        planilha.cell(linha, 1, rotulo)
        planilha.cell(linha, 2, valor)
        planilha.cell(linha, 2).number_format = "#,##0"

    planilha["D4"] = "Legenda"
    planilha["E4"] = "Uso"
    legenda = [
        ("Provável WhatsApp", "Celular brasileiro com DDD"),
        ("Verificar", "Fixo ou formato que exige conferência"),
        ("Sem telefone", "Buscar outro canal ou completar cadastro"),
    ]
    for linha, (status, uso) in enumerate(legenda, start=5):
        planilha.cell(linha, 4, status)
        planilha.cell(linha, 5, uso)
    planilha["D5"].fill = PatternFill("solid", fgColor=COR_VERDE_CLARO)
    planilha["D6"].fill = PatternFill("solid", fgColor=COR_AMARELO_CLARO)
    planilha["D7"].fill = PatternFill("solid", fgColor=COR_VERMELHO_CLARO)

    planilha.merge_cells("G4:J4")
    planilha["G4"] = "Recomendação de primeiro lote"
    planilha.merge_cells("G5:J8")
    planilha["G5"] = (
        "Comece com 20 leads da aba top_50_primeira_abordagem. Revise Google e Instagram, "
        "envie as mensagens manualmente e registre cada retorno antes de ampliar o volume."
    )
    planilha["G5"].alignment = Alignment(wrap_text=True, vertical="top")
    planilha["G5"].fill = PatternFill("solid", fgColor=COR_VERDE_CLARO)

    secoes = [
        ("Total por cidade", 1, mapa.get("cidade"), Counter(valor_texto(obter(item, mapa, "cidade")) or "Não informado" for item in registros)),
        ("Total por segmento", 4, mapa.get("segmento"), Counter(valor_texto(obter(item, mapa, "segmento")) or "Não informado" for item in registros)),
        ("Total por abordagem", 7, "abordagem_tipo", Counter(item["abordagem_tipo"] for item in registros)),
        ("Total por classificação", 10, mapa.get("classificacao"), Counter(valor_texto(obter(item, mapa, "classificacao")) or "Não informado" for item in registros)),
    ]
    inicio_tabelas = 12
    for titulo, coluna_inicio, coluna_base, contador in secoes:
        planilha.cell(inicio_tabelas, coluna_inicio, titulo)
        planilha.cell(inicio_tabelas, coluna_inicio + 1, "Quantidade")
        if coluna_base and coluna_base in indice:
            letra_base = get_column_letter(indice[coluna_base])
        elif coluna_base == "abordagem_tipo":
            letra_base = get_column_letter(indice["abordagem_tipo"])
        else:
            letra_base = None
        for deslocamento, (rotulo, quantidade) in enumerate(
            sorted(contador.items(), key=lambda item: (-item[1], item[0])), start=1
        ):
            linha = inicio_tabelas + deslocamento
            planilha.cell(linha, coluna_inicio, rotulo)
            if letra_base:
                rotulo_formula = str(rotulo).replace('"', '""')
                planilha.cell(
                    linha,
                    coluna_inicio + 1,
                    f'=COUNTIF({referencia_base}!${letra_base}$2:${letra_base}${ultima_linha},"{rotulo_formula}")',
                )
            else:
                planilha.cell(linha, coluna_inicio + 1, quantidade)
            planilha.cell(linha, coluna_inicio + 1).number_format = "#,##0"

    agregados: dict[str, dict[str, float]] = defaultdict(
        lambda: {"total": 0, "score": 0, "provavel": 0, "quente": 0}
    )
    for item in registros:
        segmento = valor_texto(obter(item, mapa, "segmento")) or "Não informado"
        agregado = agregados[segmento]
        agregado["total"] += 1
        agregado["score"] += item["_score_numero"]
        agregado["provavel"] += item["whatsapp_status"] == "Provável WhatsApp"
        agregado["quente"] += item["_classificacao_rank"] == 3
    top_segmentos = sorted(
        agregados.items(),
        key=lambda item: (
            -item[1]["quente"],
            -(item[1]["score"] / item[1]["total"] if item[1]["total"] else 0),
            -item[1]["provavel"],
            -item[1]["total"],
        ),
    )[:10]
    planilha["M12"] = "Top 10 segmentos mais promissores"
    cab_top = ["Segmento", "Leads", "Provável WhatsApp", "Quentes", "Score médio"]
    for coluna, titulo in enumerate(cab_top, start=13):
        planilha.cell(13, coluna, titulo)
    for linha, (segmento, dados) in enumerate(top_segmentos, start=14):
        planilha.cell(linha, 13, segmento)
        planilha.cell(linha, 14, int(dados["total"]))
        planilha.cell(linha, 15, int(dados["provavel"]))
        planilha.cell(linha, 16, int(dados["quente"]))
        planilha.cell(linha, 17, dados["score"] / dados["total"] if dados["total"] else 0)
        planilha.cell(linha, 17).number_format = "0.0"

    planilha["M25"] = "Arquivo de origem"
    planilha["N25"] = str(arquivo_origem.relative_to(BASE_DIR))
    planilha["M26"] = "Gerado em"
    planilha["N26"] = datetime.now()
    planilha["N26"].number_format = "dd/mm/yyyy hh:mm"

    cabecalhos_resumo = ["A4:B4", "D4:E4", "G4:J4", "A12:B12", "D12:E12", "G12:H12", "J12:K12", "M12:Q13"]
    for intervalo in cabecalhos_resumo:
        for linha in planilha[intervalo]:
            for celula in linha:
                celula.fill = PatternFill("solid", fgColor=COR_AZUL)
                celula.font = Font(color=COR_BRANCO, bold=True)
                celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    larguras = {
        "A": 33, "B": 14, "C": 3, "D": 24, "E": 42, "F": 3,
        "G": 24, "H": 20, "I": 20, "J": 26, "K": 14, "L": 3,
        "M": 34, "N": 16, "O": 20, "P": 14, "Q": 14,
    }
    for coluna, largura in larguras.items():
        planilha.column_dimensions[coluna].width = largura
    for linha in planilha.iter_rows(min_row=4, max_row=planilha.max_row, min_col=1, max_col=17):
        for celula in linha:
            celula.alignment = Alignment(vertical="top", wrap_text=True)


def gerar_excel(
    destino: Path,
    arquivo_origem: Path,
    cabecalho_original: list[str],
    registros: list[dict[str, Any]],
    mapa: dict[str, str | None],
) -> dict[str, int]:
    cabecalho = list(cabecalho_original)
    for coluna in COLUNAS_GERADAS:
        if coluna not in cabecalho:
            cabecalho.append(coluna)

    ordenados = ordenar_base(registros)
    top_50 = selecionar_top_50(ordenados, mapa)
    sem_telefone = [item for item in ordenados if item["whatsapp_status"] == "Sem telefone"]
    verificar = [item for item in ordenados if item["whatsapp_status"] == "Verificar"]

    workbook = Workbook()
    workbook.remove(workbook.active)
    escrever_aba_dados(
        workbook,
        "base_abordagem_completa",
        cabecalho,
        ordenados,
        "TabelaBaseAbordagem",
    )
    escrever_aba_dados(
        workbook,
        "top_50_primeira_abordagem",
        cabecalho,
        top_50,
        "TabelaTop50",
    )
    escrever_aba_dados(
        workbook,
        "leads_sem_telefone",
        cabecalho,
        sem_telefone,
        "TabelaSemTelefone",
    )
    escrever_aba_dados(
        workbook,
        "leads_verificar_telefone",
        cabecalho,
        verificar,
        "TabelaVerificarTelefone",
    )
    criar_resumo(workbook, cabecalho, ordenados, mapa, arquivo_origem, top_50)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    destino.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destino)

    return {
        "total": len(ordenados),
        "com_link": sum(bool(item["link_whatsapp"]) for item in ordenados),
        "provavel": sum(item["whatsapp_status"] == "Provável WhatsApp" for item in ordenados),
        "verificar": len(verificar),
        "sem_telefone": len(sem_telefone),
        "top_50": len(top_50),
    }


def executar() -> int:
    args = argumentos()
    try:
        arquivo, aba, _ = escolher_arquivo(args.arquivo)
        cabecalho, originais = ler_dados(arquivo, aba)
        mapa = mapear_colunas(cabecalho)
        registros = preparar_registros(originais, mapa)
        destino = Path(args.saida)
        if not destino.is_absolute():
            destino = BASE_DIR / destino
        totais = gerar_excel(destino, arquivo, cabecalho, registros, mapa)
    except PermissionError as erro:
        print("\nNão foi possível salvar a planilha.", file=sys.stderr)
        print("Feche o arquivo no Excel e execute o comando novamente.", file=sys.stderr)
        print(f"Detalhe: {erro}", file=sys.stderr)
        return 1
    except (FileNotFoundError, ValueError, OSError) as erro:
        print(f"\nErro: {erro}", file=sys.stderr)
        print(
            "Confira se a pasta output contém uma planilha de empresas sem site e tente novamente.",
            file=sys.stderr,
        )
        return 1

    print("\nPlanilha de abordagem criada com sucesso.")
    print(f"Quantidade total de leads processados: {totais['total']}")
    print(f"Quantidade com link de WhatsApp gerado: {totais['com_link']}")
    print(f"Quantidade com provável WhatsApp: {totais['provavel']}")
    print(f"Quantidade para verificar telefone: {totais['verificar']}")
    print(f"Quantidade sem telefone: {totais['sem_telefone']}")
    print(f"Leads na primeira abordagem: {totais['top_50']}")
    print(f"Arquivo final gerado: {destino}")
    print("Nenhuma mensagem foi enviada automaticamente.")
    return 0


if __name__ == "__main__":
    raise SystemExit(executar())

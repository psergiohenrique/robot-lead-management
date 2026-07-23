"""Busca empresas na Google Places API (New) e exporta os resultados."""

from __future__ import annotations

import argparse
import csv
import os
import sys
import time
import unicodedata
from pathlib import Path
from typing import Any

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


API_URL = "https://places.googleapis.com/v1/places:searchText"
OUTPUT_DIR = Path.cwd() / "output"
COLUNAS = [
    "nome",
    "telefone",
    "endereco",
    "google_maps",
    "avaliacao",
    "quantidade_avaliacoes",
    "site_cadastrado",
    "sem_site_cadastrado",
]
FIELD_MASK = ",".join(
    [
        "places.id",
        "places.displayName",
        "places.nationalPhoneNumber",
        "places.formattedAddress",
        "places.googleMapsUri",
        "places.rating",
        "places.userRatingCount",
        "places.websiteUri",
        "nextPageToken",
    ]
)


def argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Busca empresas por segmento e cidade usando a Google Places API (New)."
    )
    parser.add_argument("--cidade", required=True, help='Ex.: "São José do Rio Preto SP"')
    parser.add_argument("--segmento", required=True, help='Ex.: "dentista"')
    parser.add_argument("--limite", type=int, default=20, help="Máximo desejado (padrão: 20)")
    args = parser.parse_args()
    if args.limite < 1:
        parser.error("--limite deve ser maior que zero.")
    return args


def chave_normalizada(valor: str) -> str:
    sem_acentos = "".join(
        caractere
        for caractere in unicodedata.normalize("NFKD", valor)
        if not unicodedata.combining(caractere)
    )
    return " ".join(sem_acentos.casefold().split())


def buscar(api_key: str, cidade: str, segmento: str, limite: int) -> list[dict[str, Any]]:
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": FIELD_MASK,
    }
    consulta = f"{segmento} em {cidade}"
    page_token: str | None = None
    resultados: list[dict[str, Any]] = []

    while len(resultados) < limite:
        corpo: dict[str, Any] = {
            "textQuery": consulta,
            "languageCode": "pt-BR",
            "regionCode": "BR",
            "pageSize": min(20, limite - len(resultados)),
        }
        if page_token:
            corpo["pageToken"] = page_token

        try:
            resposta = requests.post(API_URL, headers=headers, json=corpo, timeout=30)
        except requests.RequestException as erro:
            raise RuntimeError(f"Falha de conexão com a API: {erro}") from erro

        if not resposta.ok:
            try:
                detalhe = resposta.json().get("error", {}).get("message", resposta.text)
            except ValueError:
                detalhe = resposta.text
            raise RuntimeError(f"Google Places API retornou HTTP {resposta.status_code}: {detalhe}")

        dados = resposta.json()
        resultados.extend(dados.get("places", []))
        page_token = dados.get("nextPageToken")
        if not page_token:
            break
        time.sleep(1)

    return resultados[:limite]


def preparar_leads(lugares: list[dict[str, Any]]) -> list[dict[str, Any]]:
    leads: list[dict[str, Any]] = []
    vistos: set[str] = set()

    for lugar in lugares:
        nome = lugar.get("displayName", {}).get("text", "")
        endereco = lugar.get("formattedAddress", "")
        telefone = lugar.get("nationalPhoneNumber", "")
        identificador = lugar.get("id") or "|".join(
            [chave_normalizada(nome), chave_normalizada(endereco), chave_normalizada(telefone)]
        )
        if identificador in vistos:
            continue
        vistos.add(identificador)

        site = lugar.get("websiteUri", "")
        leads.append(
            {
                "nome": nome,
                "telefone": telefone,
                "endereco": endereco,
                "google_maps": lugar.get("googleMapsUri", ""),
                "avaliacao": lugar.get("rating", ""),
                "quantidade_avaliacoes": lugar.get("userRatingCount", ""),
                "site_cadastrado": site,
                "sem_site_cadastrado": "SIM" if not site else "NÃO",
            }
        )
    return leads


def salvar_csv(leads: list[dict[str, Any]], caminho: Path) -> None:
    with caminho.open("w", newline="", encoding="utf-8-sig") as arquivo:
        escritor = csv.DictWriter(arquivo, fieldnames=COLUNAS, delimiter=";")
        escritor.writeheader()
        escritor.writerows(leads)


def salvar_excel(leads: list[dict[str, Any]], caminho: Path) -> None:
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Leads"
    planilha.append(COLUNAS)
    for lead in leads:
        planilha.append([lead[coluna] for coluna in COLUNAS])

    preenchimento = PatternFill("solid", fgColor="1F4E78")
    for celula in planilha[1]:
        celula.fill = preenchimento
        celula.font = Font(color="FFFFFF", bold=True)
        celula.alignment = Alignment(horizontal="center")

    larguras = [34, 20, 48, 44, 12, 22, 44, 22]
    for indice, largura in enumerate(larguras, start=1):
        planilha.column_dimensions[get_column_letter(indice)].width = largura
    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = planilha.dimensions
    planilha.sheet_view.showGridLines = False
    planilha.column_dimensions["E"].width = 12
    for linha in range(2, planilha.max_row + 1):
        planilha.cell(linha, 5).number_format = "0.0"
        planilha.cell(linha, 6).number_format = "#,##0"
        for coluna in range(1, len(COLUNAS) + 1):
            planilha.cell(linha, coluna).alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(caminho)


def main() -> int:
    args = argumentos()
    load_dotenv(Path.cwd() / ".env")
    api_key = os.getenv("GOOGLE_PLACES_API_KEY", "").strip()
    if not api_key:
        print(
            "ERRO: GOOGLE_PLACES_API_KEY não foi encontrada. "
            "Crie o arquivo .env conforme o README.md.",
            file=sys.stderr,
        )
        return 1

    print(f'Buscando até {args.limite} resultado(s) para "{args.segmento}" em "{args.cidade}"...')
    try:
        lugares = buscar(api_key, args.cidade, args.segmento, args.limite)
        leads = preparar_leads(lugares)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        salvar_csv(leads, OUTPUT_DIR / "leads.csv")
        salvar_excel(leads, OUTPUT_DIR / "leads.xlsx")
    except (RuntimeError, OSError) as erro:
        print(f"ERRO: {erro}", file=sys.stderr)
        return 1

    sem_site = sum(lead["sem_site_cadastrado"] == "SIM" for lead in leads)
    print(f"Concluído: {len(leads)} lead(s), sendo {sem_site} sem site cadastrado.")
    print(f"Arquivos: {OUTPUT_DIR / 'leads.csv'} e {OUTPUT_DIR / 'leads.xlsx'}")
    if len(lugares) < args.limite:
        print(
            "Aviso: a API devolveu menos resultados que o limite solicitado. "
            "A Text Search (New) atualmente limita uma consulta a até 60 resultados."
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

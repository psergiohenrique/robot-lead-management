"""Executa buscas de leads a partir de uma planilha de cidades e segmentos."""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import sys
import time
import unicodedata
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


API_URL = "https://places.googleapis.com/v1/places:searchText"
BASE_DIR = Path.cwd()
OUTPUT_DIR = BASE_DIR / "output"
MASTER_DIR = OUTPUT_DIR / "master"
FIELD_MASK = ",".join(
    [
        "places.id",
        "places.displayName",
        "places.formattedAddress",
        "places.nationalPhoneNumber",
        "places.websiteUri",
        "places.googleMapsUri",
        "places.rating",
        "places.userRatingCount",
        "places.businessStatus",
        "nextPageToken",
    ]
)

COLUNAS_ESPERADAS = [
    "Cidade",
    "Segmento",
    "Prioridade",
    "Região",
    "Limite Leads",
    "Oferta Principal",
    "Observação Comercial",
    "Query Base",
]

COLUNAS_SAIDA = [
    "place_id",
    "nome",
    "telefone",
    "endereco",
    "google_maps",
    "avaliacao",
    "quantidade_avaliacoes",
    "site_cadastrado",
    "sem_site_cadastrado",
    "business_status",
    "score_oportunidade",
    "classificacao_lead",
    "motivo_oportunidade",
    "cidade",
    "segmento",
    "prioridade",
    "regiao",
    "oferta_principal",
    "observacao_comercial",
    "query_base",
    "consultas_encontradas",
    "data_coleta",
]

SEGMENTOS_BONUS = {
    "dentista",
    "clinica estetica",
    "moveis planejados",
    "escritorio de arquitetura",
    "pequena industria",
    "imobiliaria",
}

ARQUIVOS_GERADOS = {
    "leads_todos.xlsx",
    "leads_sem_site.xlsx",
    "resumo_execucao.xlsx",
    "leads.xlsx",
}


def normalizar_texto(valor: Any) -> str:
    texto = "" if valor is None else str(valor).strip()
    texto = "".join(
        caractere
        for caractere in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(caractere)
    )
    return " ".join(texto.casefold().split())


def normalizar_coluna(valor: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", normalizar_texto(valor))


MAPA_COLUNAS = {normalizar_coluna(coluna): coluna for coluna in COLUNAS_ESPERADAS}


def argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Busca leads na Places API (New) a partir de uma planilha."
    )
    parser.add_argument(
        "--arquivo",
        help="Arquivo .xlsx ou .csv. Se omitido, o script procura a base automaticamente.",
    )
    parser.add_argument(
        "--prioridade",
        default="Alta",
        help='Prioridade a executar ou "Todas" (padrão: Alta).',
    )
    parser.add_argument("--cidade", help="Executa somente a cidade informada.")
    parser.add_argument("--segmento", help="Executa somente o segmento informado.")
    parser.add_argument(
        "--segmentos",
        nargs="+",
        help="Executa uma lista de segmentos. Informe cada segmento entre aspas.",
    )
    parser.add_argument(
        "--pares",
        nargs="+",
        help='Seleciona pares exatos no formato "Cidade|Segmento".',
    )
    parser.add_argument(
        "--nome-lote",
        help="Nome curto para identificar a pasta desta execução.",
    )
    parser.add_argument(
        "--max-linhas",
        type=int,
        help="Limita a quantidade de linhas da base. Útil para testes.",
    )
    args = parser.parse_args()
    if args.max_linhas is not None and args.max_linhas < 1:
        parser.error("--max-linhas deve ser maior que zero.")
    return args


def mapear_cabecalho(cabecalho: Iterable[Any]) -> dict[str, int] | None:
    encontrados: dict[str, int] = {}
    for indice, nome in enumerate(cabecalho):
        canonico = MAPA_COLUNAS.get(normalizar_coluna(nome))
        if canonico:
            encontrados[canonico] = indice
    if set(encontrados) == set(COLUNAS_ESPERADAS):
        return encontrados
    return None


def ler_xlsx(caminho: Path) -> tuple[list[dict[str, Any]], str]:
    workbook = load_workbook(caminho, read_only=True, data_only=True)
    for planilha in workbook.worksheets:
        linhas = planilha.iter_rows(values_only=True)
        cabecalho = next(linhas, None)
        if not cabecalho:
            continue
        mapa = mapear_cabecalho(cabecalho)
        if not mapa:
            continue
        registros = []
        for linha in linhas:
            registro = {
                coluna: linha[indice] if indice < len(linha) else None
                for coluna, indice in mapa.items()
            }
            if any(valor not in (None, "") for valor in registro.values()):
                registros.append(registro)
        return registros, planilha.title
    raise ValueError(
        f"Nenhuma aba de {caminho.name} contém todas as colunas esperadas."
    )


def ler_csv(caminho: Path) -> tuple[list[dict[str, Any]], str]:
    with caminho.open("r", encoding="utf-8-sig", newline="") as arquivo:
        amostra = arquivo.read(4096)
        arquivo.seek(0)
        try:
            dialeto = csv.Sniffer().sniff(amostra, delimiters=";,\t,")
        except csv.Error:
            dialeto = csv.excel
            dialeto.delimiter = ";"
        leitor = csv.reader(arquivo, dialect=dialeto)
        cabecalho = next(leitor, None)
        if not cabecalho:
            raise ValueError(f"O arquivo {caminho.name} está vazio.")
        mapa = mapear_cabecalho(cabecalho)
        if not mapa:
            raise ValueError(f"O arquivo {caminho.name} não contém todas as colunas esperadas.")
        registros = []
        for linha in leitor:
            registro = {
                coluna: linha[indice] if indice < len(linha) else ""
                for coluna, indice in mapa.items()
            }
            if any(str(valor).strip() for valor in registro.values()):
                registros.append(registro)
        return registros, "CSV"


def ler_base(caminho: Path) -> tuple[list[dict[str, Any]], str]:
    if caminho.suffix.lower() == ".xlsx":
        return ler_xlsx(caminho)
    if caminho.suffix.lower() == ".csv":
        return ler_csv(caminho)
    raise ValueError("A base precisa ser um arquivo .xlsx ou .csv.")


def localizar_arquivo(explicito: str | None) -> tuple[Path, list[Path]]:
    if explicito:
        caminho = Path(explicito).expanduser()
        if not caminho.is_absolute():
            caminho = BASE_DIR / caminho
        caminho = caminho.resolve()
        if not caminho.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")
        return caminho, [caminho]

    candidatos = [
        caminho
        for caminho in BASE_DIR.iterdir()
        if caminho.is_file()
        and caminho.suffix.lower() in {".xlsx", ".csv"}
        and not caminho.name.startswith("~$")
        and caminho.name.lower() not in ARQUIVOS_GERADOS
    ]
    validos: list[Path] = []
    for candidato in candidatos:
        try:
            ler_base(candidato)
        except (OSError, ValueError):
            continue
        validos.append(candidato)
    if not validos:
        raise FileNotFoundError(
            "Nenhum .xlsx ou .csv com as colunas esperadas foi encontrado na pasta do projeto."
        )

    def pontuacao(caminho: Path) -> tuple[int, float]:
        nome = normalizar_texto(caminho.stem)
        pontos = sum(palavra in nome for palavra in ("base", "cidade", "prioridade", "lead"))
        return pontos, caminho.stat().st_mtime

    validos.sort(key=pontuacao, reverse=True)
    return validos[0], validos


def inteiro_positivo(valor: Any, padrao: int = 20) -> int:
    try:
        numero = int(float(str(valor).replace(",", ".")))
    except (TypeError, ValueError):
        return padrao
    return max(1, numero)


def selecionar_linhas(registros: list[dict[str, Any]], args: argparse.Namespace) -> list[dict[str, Any]]:
    selecionados = []
    todas_prioridades = normalizar_texto(args.prioridade) == "todas"
    segmentos_pedidos = {
        normalizar_texto(segmento)
        for segmento in ([args.segmento] if args.segmento else []) + (args.segmentos or [])
    }
    pares_pedidos = set()
    for par in args.pares or []:
        partes = par.split("|", 1)
        if len(partes) != 2 or not all(parte.strip() for parte in partes):
            raise ValueError(f'Par inválido: "{par}". Use "Cidade|Segmento".')
        pares_pedidos.add((normalizar_texto(partes[0]), normalizar_texto(partes[1])))
    for registro in registros:
        if not todas_prioridades and normalizar_texto(registro["Prioridade"]) != normalizar_texto(
            args.prioridade
        ):
            continue
        par_registro = (
            normalizar_texto(registro["Cidade"]),
            normalizar_texto(registro["Segmento"]),
        )
        if pares_pedidos:
            if par_registro not in pares_pedidos:
                continue
        else:
            if args.cidade and normalizar_texto(registro["Cidade"]) != normalizar_texto(args.cidade):
                continue
            if segmentos_pedidos and normalizar_texto(registro["Segmento"]) not in segmentos_pedidos:
                continue
        registro = dict(registro)
        registro["Limite Leads"] = inteiro_positivo(registro["Limite Leads"])
        selecionados.append(registro)
    if args.max_linhas:
        selecionados = selecionados[: args.max_linhas]
    return selecionados


def consultas_da_linha(registro: dict[str, Any]) -> list[str]:
    cidade = str(registro["Cidade"]).strip()
    segmento = str(registro["Segmento"]).strip()
    candidatos = [
        str(registro["Query Base"] or "").strip(),
        f"{segmento} em {cidade}",
        f"{segmento} perto de {cidade}",
        f"melhor {segmento} em {cidade}",
        f"{segmento} {cidade}",
    ]
    consultas: list[str] = []
    vistos: set[str] = set()
    for consulta in candidatos:
        chave = normalizar_texto(consulta)
        if consulta and chave not in vistos:
            consultas.append(consulta)
            vistos.add(chave)
    return consultas


def mensagem_erro_api(resposta: requests.Response) -> str:
    try:
        detalhe = resposta.json().get("error", {}).get("message", resposta.text)
    except ValueError:
        detalhe = resposta.text
    explicacoes = {
        400: "requisição inválida",
        401: "chave ou autenticação inválida",
        403: "API desativada, faturamento ausente ou restrição incorreta na chave",
        429: "limite/cota da API atingido",
    }
    simples = explicacoes.get(resposta.status_code, "erro retornado pela API")
    return f"HTTP {resposta.status_code} ({simples}): {detalhe}"


def buscar_consulta(
    sessao: requests.Session,
    api_key: str,
    consulta: str,
    maximo: int,
) -> list[dict[str, Any]]:
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": FIELD_MASK,
    }
    lugares: list[dict[str, Any]] = []
    token: str | None = None
    while len(lugares) < maximo:
        corpo: dict[str, Any] = {
            "textQuery": consulta,
            "languageCode": "pt-BR",
            "regionCode": "BR",
            "pageSize": min(20, maximo - len(lugares)),
        }
        if token:
            corpo["pageToken"] = token
        try:
            resposta = sessao.post(API_URL, headers=headers, json=corpo, timeout=30)
        except requests.RequestException as erro:
            raise RuntimeError(f"falha de conexão: {erro}") from erro
        if not resposta.ok:
            raise RuntimeError(mensagem_erro_api(resposta))
        dados = resposta.json()
        lugares.extend(dados.get("places", []))
        token = dados.get("nextPageToken")
        if not token:
            break
        time.sleep(1)
    return lugares[:maximo]


def padronizar_telefone(valor: Any) -> str:
    original = "" if valor is None else str(valor).strip()
    digitos = re.sub(r"\D", "", original)
    if digitos.startswith("55") and len(digitos) in {12, 13}:
        digitos = digitos[2:]
    if len(digitos) == 11:
        return f"({digitos[:2]}) {digitos[2:7]}-{digitos[7:]}"
    if len(digitos) == 10:
        return f"({digitos[:2]}) {digitos[2:6]}-{digitos[6:]}"
    return original


def pontuar(lugar: dict[str, Any], registro: dict[str, Any]) -> tuple[int, str, str]:
    site = str(lugar.get("websiteUri", "")).strip()
    avaliacao = float(lugar.get("rating") or 0)
    quantidade = int(lugar.get("userRatingCount") or 0)
    telefone = str(lugar.get("nationalPhoneNumber", "")).strip()
    prioridade_alta = normalizar_texto(registro["Prioridade"]) == "alta"
    segmento_bonus = normalizar_texto(registro["Segmento"]) in SEGMENTOS_BONUS

    score = 0
    motivos = []
    if not site:
        score += 3
        motivos.append("sem site cadastrado")
    if avaliacao >= 4:
        score += 2
        motivos.append("boa avaliação")
    if quantidade >= 20:
        score += 2
        motivos.append("boa quantidade de avaliações")
    if telefone:
        score += 1
        motivos.append("telefone disponível")
    if prioridade_alta:
        score += 1
        motivos.append("prioridade Alta")
    if segmento_bonus:
        score += 1
        motivos.append("segmento estratégico")

    classificacao = "Quente" if score >= 8 else "Morno" if score >= 5 else "Frio"
    motivo = ", ".join(motivos) if motivos else "poucos sinais comerciais disponíveis"
    return score, classificacao, motivo


def criar_lead(
    lugar: dict[str, Any], registro: dict[str, Any], consulta: str, data_coleta: str
) -> dict[str, Any]:
    site = str(lugar.get("websiteUri", "")).strip()
    score, classificacao, motivo = pontuar(lugar, registro)
    return {
        "place_id": lugar.get("id", ""),
        "nome": lugar.get("displayName", {}).get("text", ""),
        "telefone": padronizar_telefone(lugar.get("nationalPhoneNumber", "")),
        "endereco": lugar.get("formattedAddress", ""),
        "google_maps": lugar.get("googleMapsUri", ""),
        "avaliacao": lugar.get("rating", ""),
        "quantidade_avaliacoes": lugar.get("userRatingCount", ""),
        "site_cadastrado": site,
        "sem_site_cadastrado": "SIM" if not site else "NÃO",
        "business_status": lugar.get("businessStatus", ""),
        "score_oportunidade": score,
        "classificacao_lead": classificacao,
        "motivo_oportunidade": motivo,
        "cidade": registro["Cidade"],
        "segmento": registro["Segmento"],
        "prioridade": registro["Prioridade"],
        "regiao": registro["Região"],
        "oferta_principal": registro["Oferta Principal"],
        "observacao_comercial": registro["Observação Comercial"],
        "query_base": registro["Query Base"],
        "consultas_encontradas": consulta,
        "data_coleta": data_coleta,
    }


def ordenar_leads(leads: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        leads,
        key=lambda lead: (
            lead["sem_site_cadastrado"] != "SIM",
            -int(lead["score_oportunidade"]),
            normalizar_texto(lead["prioridade"]) != "alta",
            -int(lead["quantidade_avaliacoes"] or 0),
        ),
    )


def salvar_csv(leads: list[dict[str, Any]], caminho: Path) -> None:
    with caminho.open("w", encoding="utf-8-sig", newline="") as arquivo:
        escritor = csv.DictWriter(arquivo, fieldnames=COLUNAS_SAIDA, delimiter=";")
        escritor.writeheader()
        escritor.writerows(leads)


def estilizar_planilha(planilha: Any, larguras: list[int] | None = None) -> None:
    cor = PatternFill("solid", fgColor="1F4E78")
    for celula in planilha[1]:
        celula.fill = cor
        celula.font = Font(color="FFFFFF", bold=True)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = planilha.dimensions
    planilha.sheet_view.showGridLines = False
    if larguras:
        for indice, largura in enumerate(larguras, start=1):
            planilha.column_dimensions[get_column_letter(indice)].width = largura
    else:
        for indice in range(1, planilha.max_column + 1):
            maior = max(
                (len(str(planilha.cell(linha, indice).value or "")) for linha in range(1, planilha.max_row + 1)),
                default=10,
            )
            planilha.column_dimensions[get_column_letter(indice)].width = min(max(maior + 2, 12), 45)
    for linha in planilha.iter_rows(min_row=2):
        for celula in linha:
            celula.alignment = Alignment(vertical="top", wrap_text=True)


def salvar_excel_leads(leads: list[dict[str, Any]], caminho: Path) -> None:
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Leads"
    planilha.append(COLUNAS_SAIDA)
    for lead in leads:
        planilha.append([lead[coluna] for coluna in COLUNAS_SAIDA])
    larguras = [24, 34, 20, 45, 40, 12, 20, 38, 18, 20, 20, 20, 45, 28, 25, 15, 18, 35, 45, 38, 50, 18]
    estilizar_planilha(planilha, larguras)
    for linha in range(2, planilha.max_row + 1):
        planilha.cell(linha, 6).number_format = "0.0"
        planilha.cell(linha, 7).number_format = "#,##0"
        planilha.cell(linha, 11).number_format = "0"
    workbook.save(caminho)


def nome_seguro(valor: str) -> str:
    nome = normalizar_texto(valor)
    nome = re.sub(r"[^a-z0-9]+", "-", nome).strip("-")
    return nome or "execucao"


def criar_pasta_lote(nome_lote: str | None, registros: list[dict[str, Any]]) -> Path:
    if nome_lote:
        identificador = nome_seguro(nome_lote)
    else:
        cidades = sorted({str(registro["Cidade"]) for registro in registros})
        segmentos = sorted({str(registro["Segmento"]) for registro in registros})
        identificador = nome_seguro("-".join(cidades[:1] + segmentos[:2]))
    carimbo = datetime.now().strftime("%Y%m%d_%H%M%S")
    pasta = OUTPUT_DIR / "lotes" / f"{carimbo}_{identificador}"
    contador = 2
    while pasta.exists():
        pasta = OUTPUT_DIR / "lotes" / f"{carimbo}_{identificador}_{contador}"
        contador += 1
    pasta.mkdir(parents=True, exist_ok=False)
    return pasta


def carregar_leads_csv(caminho: Path) -> list[dict[str, Any]]:
    with caminho.open("r", encoding="utf-8-sig", newline="") as arquivo:
        leitor = csv.DictReader(arquivo, delimiter=";")
        leads = []
        for linha in leitor:
            lead = {coluna: linha.get(coluna, "") for coluna in COLUNAS_SAIDA}
            for coluna in ("quantidade_avaliacoes", "score_oportunidade"):
                try:
                    lead[coluna] = int(float(str(lead[coluna]).replace(",", ".")))
                except (TypeError, ValueError):
                    lead[coluna] = 0
            try:
                lead["avaliacao"] = float(str(lead["avaliacao"]).replace(",", "."))
            except (TypeError, ValueError):
                lead["avaliacao"] = ""
            leads.append(lead)
    return leads


def carregar_master() -> tuple[list[dict[str, Any]], str]:
    master = MASTER_DIR / "leads_master.csv"
    if master.exists():
        return carregar_leads_csv(master), master.name
    legado = OUTPUT_DIR / "leads_todos.csv"
    if legado.exists():
        return carregar_leads_csv(legado), legado.name
    return [], "base vazia"


def mesclar_consultas(atual: Any, novas: Any) -> str:
    resultado: list[str] = []
    vistos: set[str] = set()
    for bloco in (str(atual or ""), str(novas or "")):
        for consulta in bloco.split(" | "):
            consulta = consulta.strip()
            chave = normalizar_texto(consulta)
            if consulta and chave not in vistos:
                resultado.append(consulta)
                vistos.add(chave)
    return " | ".join(resultado)


def mesclar_master(
    existentes: list[dict[str, Any]], novos: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], int, int]:
    por_id = {str(lead.get("place_id", "")): dict(lead) for lead in existentes if lead.get("place_id")}
    adicionados = 0
    repetidos = 0
    for lead in novos:
        place_id = str(lead.get("place_id", ""))
        if not place_id:
            continue
        if place_id in por_id:
            repetidos += 1
            por_id[place_id]["consultas_encontradas"] = mesclar_consultas(
                por_id[place_id].get("consultas_encontradas"), lead.get("consultas_encontradas")
            )
        else:
            por_id[place_id] = dict(lead)
            adicionados += 1
    return ordenar_leads(por_id.values()), adicionados, repetidos


def salvar_master(leads: list[dict[str, Any]]) -> None:
    MASTER_DIR.mkdir(parents=True, exist_ok=True)
    historico = MASTER_DIR / "historico"
    historico.mkdir(parents=True, exist_ok=True)
    carimbo = datetime.now().strftime("%Y%m%d_%H%M%S")
    for nome in ("leads_master.csv", "leads_master.xlsx"):
        atual = MASTER_DIR / nome
        if atual.exists():
            shutil.copy2(atual, historico / f"{carimbo}_{nome}")

    sem_site = [lead for lead in leads if lead["sem_site_cadastrado"] == "SIM"]
    salvar_csv(leads, MASTER_DIR / "leads_master.csv")
    salvar_excel_leads(leads, MASTER_DIR / "leads_master.xlsx")
    salvar_csv(sem_site, MASTER_DIR / "leads_master_sem_site.csv")
    salvar_excel_leads(sem_site, MASTER_DIR / "leads_master_sem_site.xlsx")


def adicionar_tabela_resumo(workbook: Workbook, titulo: str, contador: Counter[str]) -> None:
    planilha = workbook.create_sheet(titulo[:31])
    planilha.append([titulo.replace("Por ", ""), "Quantidade"])
    for chave, quantidade in sorted(contador.items(), key=lambda item: (-item[1], item[0])):
        planilha.append([chave or "Não informado", quantidade])
    estilizar_planilha(planilha, [35, 15])
    for linha in range(2, planilha.max_row + 1):
        planilha.cell(linha, 2).number_format = "#,##0"


def salvar_resumo(
    leads: list[dict[str, Any]],
    buscas: list[dict[str, Any]],
    erros: list[dict[str, Any]],
    caminho: Path,
    arquivo_base: Path,
    linhas_executadas: int,
) -> None:
    workbook = Workbook()
    resumo = workbook.active
    resumo.title = "Resumo"
    sem_site = sum(lead["sem_site_cadastrado"] == "SIM" for lead in leads)
    resumo.append(["Indicador", "Valor"])
    resumo.append(["Arquivo base", arquivo_base.name])
    resumo.append(["Linhas da base executadas", linhas_executadas])
    resumo.append(["Buscas realizadas", len(buscas)])
    resumo.append(["Total de empresas encontradas", len(leads)])
    resumo.append(["Empresas sem site cadastrado", sem_site])
    resumo.append(["Erros por busca", len(erros)])
    resumo.append(["Gerado em", datetime.now()])
    estilizar_planilha(resumo, [36, 30])
    resumo["B8"].number_format = "yyyy-mm-dd hh:mm"

    adicionar_tabela_resumo(workbook, "Por Cidade", Counter(str(lead["cidade"]) for lead in leads))
    adicionar_tabela_resumo(workbook, "Por Segmento", Counter(str(lead["segmento"]) for lead in leads))
    adicionar_tabela_resumo(workbook, "Por Prioridade", Counter(str(lead["prioridade"]) for lead in leads))
    adicionar_tabela_resumo(
        workbook, "Por Classificacao", Counter(str(lead["classificacao_lead"]) for lead in leads)
    )

    aba_buscas = workbook.create_sheet("Buscas")
    colunas_buscas = ["cidade", "segmento", "consulta", "status", "resultados_api", "novos_na_linha"]
    aba_buscas.append(colunas_buscas)
    for busca in buscas:
        aba_buscas.append([busca.get(coluna, "") for coluna in colunas_buscas])
    estilizar_planilha(aba_buscas, [28, 28, 48, 15, 18, 18])

    aba_erros = workbook.create_sheet("Erros por Busca")
    colunas_erros = ["cidade", "segmento", "consulta", "erro"]
    aba_erros.append(colunas_erros)
    for erro in erros:
        aba_erros.append([erro.get(coluna, "") for coluna in colunas_erros])
    estilizar_planilha(aba_erros, [28, 28, 48, 70])
    workbook.save(caminho)


def executar(registros: list[dict[str, Any]], api_key: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    leads_por_id: dict[str, dict[str, Any]] = {}
    buscas: list[dict[str, Any]] = []
    erros: list[dict[str, Any]] = []
    data_coleta = datetime.now().astimezone().isoformat(timespec="seconds")

    with requests.Session() as sessao:
        for numero, registro in enumerate(registros, start=1):
            cidade = str(registro["Cidade"]).strip()
            segmento = str(registro["Segmento"]).strip()
            limite = int(registro["Limite Leads"])
            print(f"\n[{numero}/{len(registros)}] {segmento} — {cidade} (limite: {limite})")
            ids_da_linha: set[str] = set()

            for consulta in consultas_da_linha(registro):
                if len(ids_da_linha) >= limite:
                    break
                restante = limite - len(ids_da_linha)
                print(f'  Buscando: "{consulta}"')
                try:
                    lugares = buscar_consulta(sessao, api_key, consulta, restante)
                except RuntimeError as erro:
                    mensagem = str(erro)
                    print(f"  ERRO: {mensagem}")
                    item_erro = {
                        "cidade": cidade,
                        "segmento": segmento,
                        "consulta": consulta,
                        "erro": mensagem,
                    }
                    erros.append(item_erro)
                    buscas.append({**item_erro, "status": "ERRO", "resultados_api": 0, "novos_na_linha": 0})
                    continue

                novos_na_linha = 0
                for lugar in lugares:
                    place_id = str(lugar.get("id", "")).strip()
                    if not place_id:
                        continue
                    status = lugar.get("businessStatus")
                    if status and status != "OPERATIONAL":
                        continue
                    if place_id not in ids_da_linha:
                        ids_da_linha.add(place_id)
                        novos_na_linha += 1
                    if place_id in leads_por_id:
                        consultas_atuais = leads_por_id[place_id]["consultas_encontradas"].split(" | ")
                        if consulta not in consultas_atuais:
                            consultas_atuais.append(consulta)
                            leads_por_id[place_id]["consultas_encontradas"] = " | ".join(consultas_atuais)
                    else:
                        leads_por_id[place_id] = criar_lead(lugar, registro, consulta, data_coleta)

                sem_site_linha = sum(
                    lead["sem_site_cadastrado"] == "SIM"
                    and lead["place_id"] in ids_da_linha
                    for lead in leads_por_id.values()
                )
                print(
                    f"  Retornados: {len(lugares)} | únicos nesta linha: {len(ids_da_linha)} "
                    f"| sem site até agora: {sem_site_linha}"
                )
                buscas.append(
                    {
                        "cidade": cidade,
                        "segmento": segmento,
                        "consulta": consulta,
                        "status": "OK",
                        "resultados_api": len(lugares),
                        "novos_na_linha": novos_na_linha,
                    }
                )

            print(f"  Total único da linha: {len(ids_da_linha)}")

    return ordenar_leads(leads_por_id.values()), buscas, erros


def main() -> int:
    args = argumentos()
    load_dotenv(BASE_DIR / ".env")
    api_key = os.getenv("GOOGLE_PLACES_API_KEY", "").strip()
    if not api_key or api_key == "sua_chave_aqui":
        print("ERRO: preencha GOOGLE_PLACES_API_KEY no arquivo .env.", file=sys.stderr)
        return 1

    try:
        arquivo, candidatos = localizar_arquivo(args.arquivo)
        registros, aba = ler_base(arquivo)
        selecionados = selecionar_linhas(registros, args)
    except (FileNotFoundError, OSError, ValueError) as erro:
        print(f"ERRO: {erro}", file=sys.stderr)
        return 1

    if len(candidatos) > 1:
        print("Bases válidas encontradas:")
        for candidato in candidatos:
            marcador = " (selecionada)" if candidato == arquivo else ""
            print(f"  - {candidato.name}{marcador}")
    if not selecionados:
        print("ERRO: nenhuma linha corresponde aos filtros informados.", file=sys.stderr)
        return 1

    print(f"Base: {arquivo.name} | aba: {aba} | linhas válidas: {len(registros)}")
    print(f"Linhas selecionadas para esta execução: {len(selecionados)}")
    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        pasta_lote = criar_pasta_lote(args.nome_lote, selecionados)
    except OSError as erro:
        print(f"ERRO ao criar a pasta do lote: {erro}", file=sys.stderr)
        return 1
    print(f"Pasta exclusiva deste lote: {pasta_lote}")

    leads, buscas, erros = executar(selecionados, api_key)
    sem_site = [lead for lead in leads if lead["sem_site_cadastrado"] == "SIM"]
    try:
        salvar_csv(leads, pasta_lote / "leads_todos.csv")
        salvar_excel_leads(leads, pasta_lote / "leads_todos.xlsx")
        salvar_csv(sem_site, pasta_lote / "leads_sem_site.csv")
        salvar_excel_leads(sem_site, pasta_lote / "leads_sem_site.xlsx")
        salvar_resumo(
            leads,
            buscas,
            erros,
            pasta_lote / "resumo_execucao.xlsx",
            arquivo,
            len(selecionados),
        )
        master_anterior, origem_master = carregar_master()
        master, adicionados_master, repetidos_master = mesclar_master(master_anterior, leads)
        salvar_master(master)
    except OSError as erro:
        print(
            f"ERRO ao salvar arquivos: {erro}. Feche os arquivos no Excel e tente novamente.",
            file=sys.stderr,
        )
        return 1

    print("\nExecução concluída.")
    print(f"Empresas únicas: {len(leads)}")
    print(f"Empresas sem site: {len(sem_site)}")
    print(f"Buscas com erro: {len(erros)}")
    print(f"Base master anterior: {origem_master} ({len(master_anterior)} leads)")
    print(f"Novos adicionados à master: {adicionados_master}")
    print(f"Já existentes na master: {repetidos_master}")
    print(f"Total atual da master: {len(master)}")
    print(f"Arquivos do lote: {pasta_lote}")
    print(f"Base master: {MASTER_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
